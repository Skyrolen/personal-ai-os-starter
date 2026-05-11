#!/usr/bin/env python3
"""Generate the restaurant back-of-house inventory + recipe + requisition template.

Output: templates/restaurant-boh-inventory.xlsx

Run:    python3 scripts/build_boh_template.py
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.stderr.write(
        "openpyxl is required. Install it with:\n    pip install openpyxl\n"
    )
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "templates" / "restaurant-boh-inventory.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
INPUT_FONT = Font(bold=True)
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
)


# -- seed data --------------------------------------------------------------

ITEMS = [
    # ItemCode, ItemName, Category, VendorUOM, PackSize, PortionUOM, PortionsPerVendorUnit, VendorPrice, Notes
    ("PORK-CHOP-6OZ", "Pork Chop 6oz",      "Protein",   "lb",    "40 lb case", "each", 2.67, 4.50,  "Bone-in. Vendor sells by the pound."),
    ("CHIX-BRST-8OZ", "Chicken Breast 8oz", "Protein",   "case",  "40 ct case", "each", 40,   95.00, "Boneless skinless."),
    ("ONION-YEL",     "Yellow Onion",       "Produce",   "lb",    "50 lb sack", "cup",  3.0,  0.95,  "Approx 3 cups diced per lb."),
    ("CREAM-HVY",     "Heavy Cream",        "Dairy",     "qt",    "1 qt",       "cup",  4,    7.25,  ""),
    ("SALT-KOSHER",   "Kosher Salt",        "Dry",       "box",   "3 lb box",   "tsp",  864,  4.10,  "Approx 288 tsp per lb."),
    ("OIL-OLIVE",     "Olive Oil",          "Dry",       "gal",   "1 gal",      "tbsp", 256,  32.00, ""),
    ("POT-RUSSET",    "Russet Potato",      "Produce",   "case",  "50 lb case", "each", 100,  22.00, "~2 potatoes per lb."),
    ("BUTTER",        "Butter Unsalted",    "Dairy",     "lb",    "1 lb block", "tbsp", 32,   5.50,  ""),
    ("GARLIC",        "Garlic Fresh",       "Produce",   "lb",    "5 lb bag",   "clove", 50,  6.00,  ""),
    ("PARSLEY",       "Parsley Flat-Leaf",  "Produce",   "bunch", "1 bunch",    "tbsp", 24,   1.75,  ""),
]

LOCATIONS = [
    # LocationCode, LocationName, Zone, PullPriority (1 = pull first)
    ("PREP",   "Prep Station",   "Prep",     1),
    ("SAUTE",  "Sauté Line",     "Hot Line", 2),
    ("WALKIN", "Walk-in Cooler", "Cold",     3),
    ("DRY",    "Dry Storage",    "Ambient", 4),
    ("FRZ",    "Freezer",        "Cold",    5),
]

# Inventory rows show one item across many locations (the "pivot").
INVENTORY = [
    # ItemCode, LocationCode, ParLevel, OnHandPortions
    ("PORK-CHOP-6OZ",  "WALKIN", 60,  40),
    ("PORK-CHOP-6OZ",  "FRZ",    120, 120),
    ("PORK-CHOP-6OZ",  "PREP",   12,  8),
    ("CHIX-BRST-8OZ",  "WALKIN", 40,  32),
    ("CHIX-BRST-8OZ",  "FRZ",    80,  80),
    ("ONION-YEL",      "WALKIN", 30,  18),
    ("ONION-YEL",      "PREP",   6,   4),
    ("CREAM-HVY",      "WALKIN", 12,  6),
    ("SALT-KOSHER",    "DRY",    500, 720),
    ("OIL-OLIVE",      "DRY",    128, 96),
    ("POT-RUSSET",     "DRY",    60,  48),
    ("POT-RUSSET",     "PREP",   12,  10),
    ("BUTTER",         "WALKIN", 16,  12),
    ("BUTTER",         "SAUTE",  4,   3),
    ("GARLIC",         "WALKIN", 20,  15),
    ("PARSLEY",        "WALKIN", 6,   4),
]

RECIPES = [
    # RecipeCode, RecipeName, YieldServings, ItemCode, QtyPerBatch, QtyUOM
    ("PORK01", "Pan-Seared Pork Chop", 1, "PORK-CHOP-6OZ", 1,    "each"),
    ("PORK01", "Pan-Seared Pork Chop", 1, "ONION-YEL",     0.25, "cup"),
    ("PORK01", "Pan-Seared Pork Chop", 1, "BUTTER",        1,    "tbsp"),
    ("PORK01", "Pan-Seared Pork Chop", 1, "SALT-KOSHER",   0.5,  "tsp"),
    ("PORK01", "Pan-Seared Pork Chop", 1, "GARLIC",        2,    "clove"),
    ("CHIX01", "Roast Chicken Plate",  1, "CHIX-BRST-8OZ", 1,    "each"),
    ("CHIX01", "Roast Chicken Plate",  1, "POT-RUSSET",    1,    "each"),
    ("CHIX01", "Roast Chicken Plate",  1, "OIL-OLIVE",     1,    "tbsp"),
    ("CHIX01", "Roast Chicken Plate",  1, "BUTTER",        1,    "tbsp"),
    ("CHIX01", "Roast Chicken Plate",  1, "PARSLEY",       1,    "tbsp"),
    ("CHIX01", "Roast Chicken Plate",  1, "SALT-KOSHER",   0.5,  "tsp"),
]


# -- helpers ----------------------------------------------------------------

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws, headers, sample_widths=None):
    sample_widths = sample_widths or {}
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)) + 2, sample_widths.get(h, 14))
        ws.column_dimensions[get_column_letter(i)].width = width


def add_table(ws, name, start_row, end_row, ncols):
    ref = f"A{start_row}:{get_column_letter(ncols)}{end_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)


# -- sheet builders ---------------------------------------------------------

def build_items(wb):
    ws = wb.create_sheet("Items")
    headers = [
        "ItemCode", "ItemName", "Category", "VendorUOM", "PackSize",
        "PortionUOM", "PortionsPerVendorUnit", "VendorPrice", "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in ITEMS:
        ws.append(list(row))
    autosize(ws, headers, {"ItemName": 22, "PackSize": 16, "Notes": 40,
                           "PortionsPerVendorUnit": 22, "VendorPrice": 14})
    add_table(ws, "tbl_Items", 1, 1 + len(ITEMS), len(headers))


def build_locations(wb):
    ws = wb.create_sheet("Locations")
    headers = ["LocationCode", "LocationName", "Zone", "PullPriority"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in LOCATIONS:
        ws.append(list(row))
    autosize(ws, headers, {"LocationName": 22, "PullPriority": 14})
    add_table(ws, "tbl_Locations", 1, 1 + len(LOCATIONS), len(headers))


def build_inventory(wb):
    ws = wb.create_sheet("Inventory")
    headers = [
        "ItemCode", "LocationCode", "ParLevel", "OnHandPortions",
        "ItemName", "PortionUOM", "OnHandVendorUnits",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r, row in enumerate(INVENTORY, start=2):
        item_code, loc_code, par, on_hand = row
        ws.cell(row=r, column=1, value=item_code)
        ws.cell(row=r, column=2, value=loc_code)
        ws.cell(row=r, column=3, value=par)
        ws.cell(row=r, column=4, value=on_hand)
        ws.cell(
            row=r, column=5,
            value='=XLOOKUP([@ItemCode], tbl_Items[ItemCode], tbl_Items[ItemName], "MISSING")',
        )
        ws.cell(
            row=r, column=6,
            value='=XLOOKUP([@ItemCode], tbl_Items[ItemCode], tbl_Items[PortionUOM], "")',
        )
        ws.cell(
            row=r, column=7,
            value='=IFERROR([@OnHandPortions]/XLOOKUP([@ItemCode], tbl_Items[ItemCode], tbl_Items[PortionsPerVendorUnit]), 0)',
        )
    autosize(ws, headers, {"OnHandVendorUnits": 20, "OnHandPortions": 18,
                           "ItemName": 22})
    add_table(ws, "tbl_Inventory", 1, 1 + len(INVENTORY), len(headers))


def build_recipes(wb):
    ws = wb.create_sheet("Recipes")
    headers = [
        "RecipeCode", "RecipeName", "YieldServings",
        "ItemCode", "QtyPerBatch", "QtyUOM",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in RECIPES:
        ws.append(list(row))
    autosize(ws, headers, {"RecipeName": 26, "ItemCode": 18})
    add_table(ws, "tbl_Recipes", 1, 1 + len(RECIPES), len(headers))


def build_requisition(wb):
    ws = wb.create_sheet("Requisition")

    ws["A1"] = "RecipeCode:"
    ws["A2"] = "ServingsNeeded:"
    ws["A1"].font = INPUT_FONT
    ws["A2"].font = INPUT_FONT
    ws["B1"] = "PORK01"
    ws["B2"] = 20
    for ref in ("B1", "B2"):
        ws[ref].fill = INPUT_FILL
        ws[ref].font = INPUT_FONT

    dv = DataValidation(
        type="list",
        formula1="=INDIRECT(\"tbl_Recipes[RecipeCode]\")",
        allow_blank=True,
        showErrorMessage=False,
    )
    dv.add("B1")
    ws.add_data_validation(dv)

    ws["D1"] = "RecipeName:"
    ws["D1"].font = INPUT_FONT
    ws["E1"] = '=IFERROR(XLOOKUP($B$1, tbl_Recipes[RecipeCode], tbl_Recipes[RecipeName]), "")'
    ws["D2"] = "Recipe Yield:"
    ws["D2"].font = INPUT_FONT
    ws["E2"] = '=IFERROR(XLOOKUP($B$1, tbl_Recipes[RecipeCode], tbl_Recipes[YieldServings]), "")'

    header_row = 4
    headers = [
        "ItemCode", "ItemName", "TotalPortionsNeeded", "PortionUOM",
        "VendorUOM", "VendorUnitsNeeded", "VendorUnitsToOrder",
        "TotalOnHandPortions", "ShortfallPortions", "EstCost",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header(ws, header_row, len(headers))

    for r, item in enumerate(ITEMS, start=header_row + 1):
        item_code = item[0]
        ws.cell(row=r, column=1, value=item_code)
        ws.cell(
            row=r, column=2,
            value=f'=XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[ItemName], "")',
        )
        ws.cell(
            row=r, column=3,
            value=(
                f'=IFERROR(SUMIFS(tbl_Recipes[QtyPerBatch], tbl_Recipes[ItemCode], A{r}, '
                f'tbl_Recipes[RecipeCode], $B$1) * $B$2 / '
                f'XLOOKUP($B$1, tbl_Recipes[RecipeCode], tbl_Recipes[YieldServings]), 0)'
            ),
        )
        ws.cell(
            row=r, column=4,
            value=f'=XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[PortionUOM], "")',
        )
        ws.cell(
            row=r, column=5,
            value=f'=XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[VendorUOM], "")',
        )
        ws.cell(
            row=r, column=6,
            value=(
                f'=IFERROR(C{r}/XLOOKUP(A{r}, tbl_Items[ItemCode], '
                f'tbl_Items[PortionsPerVendorUnit]), 0)'
            ),
        )
        ws.cell(row=r, column=7, value=f'=IF(F{r}>0, CEILING(F{r}, 1), 0)')
        ws.cell(
            row=r, column=8,
            value=f'=SUMIFS(tbl_Inventory[OnHandPortions], tbl_Inventory[ItemCode], A{r})',
        )
        ws.cell(row=r, column=9, value=f'=MAX(0, C{r}-H{r})')
        ws.cell(
            row=r, column=10,
            value=f'=G{r}*XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[VendorPrice], 0)',
        )

    end_row = header_row + len(ITEMS)
    add_table(ws, "tbl_Requisition", header_row, end_row, len(headers))

    autosize(ws, headers, {
        "ItemName": 22, "TotalPortionsNeeded": 20, "VendorUnitsNeeded": 18,
        "VendorUnitsToOrder": 20, "TotalOnHandPortions": 20,
        "ShortfallPortions": 18, "EstCost": 12,
    })
    ws.column_dimensions["A"].width = 18

    ws["A2"].alignment = Alignment(horizontal="right")
    ws["A1"].alignment = Alignment(horizontal="right")


def build_pull_list(wb):
    ws = wb.create_sheet("PullList")
    headers = [
        "ItemCode", "ItemName", "LocationCode", "LocationName",
        "PullPriority", "PortionsAvailable", "PortionUOM",
        "PortionsNeededTotal", "AlreadyAllocatedHigherPriority",
        "PortionsToPullHere", "StillShortAfterHere",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    priority_by_loc = {code: prio for code, _name, _zone, prio in LOCATIONS}
    sorted_inventory = sorted(
        INVENTORY,
        key=lambda inv: (inv[0], priority_by_loc.get(inv[1], 999)),
    )

    for r, inv in enumerate(sorted_inventory, start=2):
        item_code, loc_code, _par, on_hand = inv
        ws.cell(row=r, column=1, value=item_code)
        ws.cell(
            row=r, column=2,
            value=f'=XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[ItemName], "")',
        )
        ws.cell(row=r, column=3, value=loc_code)
        ws.cell(
            row=r, column=4,
            value=f'=XLOOKUP(C{r}, tbl_Locations[LocationCode], tbl_Locations[LocationName], "")',
        )
        ws.cell(
            row=r, column=5,
            value=f'=XLOOKUP(C{r}, tbl_Locations[LocationCode], tbl_Locations[PullPriority], 999)',
        )
        ws.cell(row=r, column=6, value=on_hand)
        ws.cell(
            row=r, column=7,
            value=f'=XLOOKUP(A{r}, tbl_Items[ItemCode], tbl_Items[PortionUOM], "")',
        )
        ws.cell(
            row=r, column=8,
            value=(
                f'=IFERROR(XLOOKUP(A{r}, tbl_Requisition[ItemCode], '
                f'tbl_Requisition[TotalPortionsNeeded]), 0)'
            ),
        )
        ws.cell(
            row=r, column=9,
            value=(
                f'=SUMIFS(tbl_PullList[PortionsToPullHere], '
                f'tbl_PullList[ItemCode], A{r}, '
                f'tbl_PullList[PullPriority], "<"&E{r})'
            ),
        )
        ws.cell(
            row=r, column=10,
            value=f'=MIN(F{r}, MAX(0, H{r}-I{r}))',
        )
        ws.cell(
            row=r, column=11,
            value=f'=MAX(0, H{r}-I{r}-J{r})',
        )

    autosize(ws, headers, {
        "ItemName": 22, "LocationName": 18,
        "PortionsAvailable": 18, "PortionsNeededTotal": 20,
        "AlreadyAllocatedHigherPriority": 30, "PortionsToPullHere": 20,
        "StillShortAfterHere": 20,
    })
    add_table(ws, "tbl_PullList", 1, 1 + len(sorted_inventory), len(headers))


def build_readme_sheet(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("Restaurant BOH Inventory Template", True),
        ("", False),
        ("Sheets:", True),
        ("  Items        - master catalog. One row per SKU. Edit PortionsPerVendorUnit to fix unit math.", False),
        ("  Locations    - storage locations + PullPriority (1 = pull from here first).", False),
        ("  Inventory    - on-hand by item AND location. One item can have many rows here.", False),
        ("  Recipes      - long-format: one row per ingredient per recipe. Paste new rows to add a recipe.", False),
        ("  Requisition  - pick a RecipeCode in B1, ServingsNeeded in B2 - totals + vendor units appear below.", False),
        ("  PullList     - allocates the requisition across locations in PullPriority order.", False),
        ("", False),
        ("Quick test:", True),
        ("  1. Open the Requisition sheet.", False),
        ("  2. Set B1 = PORK01, B2 = 20.", False),
        ("  3. Pork chop row should show TotalPortionsNeeded = 20, VendorUnitsToOrder = 8.", False),
        ("", False),
        ("Adding a recipe:", True),
        ("  Paste new rows at the bottom of the Recipes table - one row per ingredient.", False),
        ("  Use ItemCodes that already exist in Items (or add them first).", False),
        ("", False),
        ("Vendor unit conversion:", True),
        ("  Vendors sell pork chops by the pound. The kitchen thinks in each (6oz portions).", False),
        ("  PortionsPerVendorUnit on Items is the conversion: 1 lb -> 2.67 chops.", False),
        ("  Requisition reports both PortionsNeeded (kitchen) and VendorUnitsToOrder (rounded up to a whole lb/case).", False),
        ("", False),
        ("Regenerate this file: python3 scripts/build_boh_template.py", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 110


# -- entry point ------------------------------------------------------------

def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_items(wb)
    build_locations(wb)
    build_inventory(wb)
    build_recipes(wb)
    build_requisition(wb)
    build_pull_list(wb)
    build_readme_sheet(wb)

    wb.active = wb.sheetnames.index("Requisition")
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
